"""Pupa Counter — standalone inference + Excel export.

Default pipeline: v12 CNN (fresh-trained on 99 cleaned scans with
per-pixel spatial UP weighting that lightens the penalty on the 7-8px
black scanner-border region and amplifies it on paper interior) +
matched peak-level classifier. Self-eval F1 = 99.46% with 100% precision.

Usage:
    # Count pupae in a single scan
    python pupa_counter.py path/to/scan.png

    # Batch process a directory
    python pupa_counter.py path/to/scans/

    # Custom output directory + excel file
    python pupa_counter.py scans/ --out results/ --excel my_counts.xlsx

For each scan the script:
  1. Runs the v12 CNN model on the image.
  2. Detects pupa centers via heatmap peak extraction.
  3. Splits the image at 25% and 75% horizontal lines.
     Convention: rank 0% = pupa at IMAGE BOTTOM, rank 100% = pupa at IMAGE TOP.
       - Lower 25% of image  = rank 0-25% band
       - Middle 50% of image = rank 25-75% band
       - Upper 25% of image  = rank 75-100% band
  4. Draws detections + split lines; saves annotated PNG to output/.
  5. Appends a row to the Excel results file
     (created automatically on first run).
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np
import torch
import torch.nn as nn
from skimage.feature import peak_local_max

# ----------------------------------------------------------------------------
# Model definition (TinyUNet — must match training architecture)
# ----------------------------------------------------------------------------

class TinyUNet(nn.Module):
    """Small U-Net for heatmap regression (466K params)."""

    def __init__(self, in_ch: int = 3, out_ch: int = 1, base: int = 32):
        super().__init__()
        self.enc1 = nn.Sequential(
            nn.Conv2d(in_ch, base, 3, padding=1), nn.ReLU(),
            nn.Conv2d(base, base, 3, padding=1), nn.ReLU(),
        )
        self.enc2 = nn.Sequential(
            nn.Conv2d(base, base * 2, 3, padding=1), nn.ReLU(),
            nn.Conv2d(base * 2, base * 2, 3, padding=1), nn.ReLU(),
        )
        self.enc3 = nn.Sequential(
            nn.Conv2d(base * 2, base * 4, 3, padding=1), nn.ReLU(),
            nn.Conv2d(base * 4, base * 4, 3, padding=1), nn.ReLU(),
        )
        self.pool = nn.MaxPool2d(2)
        self.up2 = nn.ConvTranspose2d(base * 4, base * 2, 2, stride=2)
        self.dec2 = nn.Sequential(
            nn.Conv2d(base * 4, base * 2, 3, padding=1), nn.ReLU(),
            nn.Conv2d(base * 2, base * 2, 3, padding=1), nn.ReLU(),
        )
        self.up1 = nn.ConvTranspose2d(base * 2, base, 2, stride=2)
        self.dec1 = nn.Sequential(
            nn.Conv2d(base * 2, base, 3, padding=1), nn.ReLU(),
            nn.Conv2d(base, base, 3, padding=1), nn.ReLU(),
        )
        self.out = nn.Conv2d(base, out_ch, 1)

    def forward(self, x):
        e1 = self.enc1(x)
        e2 = self.enc2(self.pool(e1))
        e3 = self.enc3(self.pool(e2))
        d2 = self.dec2(torch.cat([self.up2(e3), e2], dim=1))
        d1 = self.dec1(torch.cat([self.up1(d2), e1], dim=1))
        return torch.sigmoid(self.out(d1))


# ----------------------------------------------------------------------------
# Inference helpers
# ----------------------------------------------------------------------------

PATCH_SIZE = 256
STRIDE = 192
PEAK_THRESHOLD = 0.15
PEAK_MIN_DIST = 8


def pick_device() -> torch.device:
    """Pick the fastest available inference backend.

    Order: MPS (Apple Silicon) > CUDA (NVIDIA) > XPU (Intel GPU via
    oneAPI, torch 2.5+) > DirectML (AMD + other Windows GPUs, optional
    package) > CPU.

    Installing a GPU-capable torch on each platform is the user's job —
    see scripts/setup_venv.py. This function just picks what's reachable
    from the currently imported torch.
    """
    if torch.backends.mps.is_available():
        return torch.device("mps")
    if torch.cuda.is_available():
        return torch.device("cuda")
    if hasattr(torch, "xpu") and torch.xpu.is_available():
        return torch.device("xpu")
    try:
        import torch_directml  # type: ignore
        return torch_directml.device()
    except ImportError:
        pass
    return torch.device("cpu")


def device_description(device: torch.device) -> str:
    """Human label for the Settings UI / CLI startup banner."""
    t = device.type
    if t == "cuda":
        try: return f"CUDA · {torch.cuda.get_device_name(0)}"
        except Exception: return "CUDA"
    if t == "mps":
        return "Apple Silicon (MPS)"
    if t == "xpu":
        try: return f"Intel XPU · {torch.xpu.get_device_name(0)}"
        except Exception: return "Intel XPU"
    if t == "privateuseone":   # torch_directml device type
        return "DirectML (Windows)"
    return "CPU"


def predict_heatmap(model: TinyUNet, img_rgb: np.ndarray, device: torch.device,
                    patch: int = PATCH_SIZE, stride: int = STRIDE) -> np.ndarray:
    """Tile the full scan, run the model per tile, average overlapping regions."""
    h, w = img_rgb.shape[:2]
    heat = np.zeros((h, w), dtype=np.float32)
    count = np.zeros((h, w), dtype=np.float32)
    model.eval()
    with torch.no_grad():
        for y0 in range(0, h, stride):
            for x0 in range(0, w, stride):
                y1 = min(h, y0 + patch)
                x1 = min(w, x0 + patch)
                if y1 - y0 < 64 or x1 - x0 < 64:
                    continue
                ph = patch - (y1 - y0)
                pw = patch - (x1 - x0)
                tile = img_rgb[y0:y1, x0:x1]
                if ph > 0 or pw > 0:
                    tile = np.pad(tile, ((0, ph), (0, pw), (0, 0)), mode="reflect")
                x_tensor = torch.from_numpy(tile.astype(np.float32) / 255.0) \
                    .permute(2, 0, 1).unsqueeze(0).to(device)
                pred = model(x_tensor).squeeze().cpu().numpy()
                ay0, ax0 = y0, x0
                ay1 = min(h, y0 + patch)
                ax1 = min(w, x0 + patch)
                h_r, w_r = ay1 - ay0, ax1 - ax0
                heat[ay0:ay1, ax0:ax1] += pred[:h_r, :w_r]
                count[ay0:ay1, ax0:ax1] += 1
    return heat / np.maximum(count, 1)


def extract_peaks(heatmap: np.ndarray,
                  threshold: float = PEAK_THRESHOLD,
                  min_dist: int = PEAK_MIN_DIST) -> list[tuple[int, int]]:
    """Extract (x, y) peak locations from heatmap."""
    coords = peak_local_max(heatmap, min_distance=min_dist, threshold_abs=threshold)
    return [(int(x), int(y)) for y, x in coords]


# ----------------------------------------------------------------------------
# Peak-level false-positive classifier (optional 2nd-stage filter)
# ----------------------------------------------------------------------------

CLASSIFIER_PATH = Path(__file__).resolve().parent / "model" / "peak_filter_clf.pkl"
CLASSIFIER_THRESHOLD = 0.60  # keep peak if P(real) >= this


def _peak_features(peaks: list[tuple[int, int]], heatmap: np.ndarray,
                   rgb: np.ndarray) -> np.ndarray:
    """11-dim feature per peak: score, local heat stats, color scores, neighbors, edge distance."""
    if len(peaks) == 0:
        return np.zeros((0, 11), dtype=np.float32)
    r = rgb[..., 0].astype(np.float32)
    g = rgb[..., 1].astype(np.float32)
    b = rgb[..., 2].astype(np.float32)
    yellow = np.clip((np.minimum(r, g) - b - 15) / 60, 0, 1)
    blue = np.clip((b - np.maximum(r, g) - 15) / 60, 0, 1)
    gray = (0.3 * r + 0.59 * g + 0.11 * b) / 255.0
    H, W = heatmap.shape

    scores = [float(heatmap[y, x]) for x, y in peaks]
    feats = np.zeros((len(peaks), 11), dtype=np.float32)
    for i, (xi, yi) in enumerate(peaks):
        si = scores[i]
        y0, y1 = max(0, yi - 4), min(H, yi + 5)
        x0, x1 = max(0, xi - 4), min(W, xi + 5)
        y0b, y1b = max(0, yi - 8), min(H, yi + 9)
        x0b, x1b = max(0, xi - 8), min(W, xi + 9)
        # Nearest neighbor
        nn_d, nn_s = 9999.0, 0.0
        for j, (xj, yj) in enumerate(peaks):
            if i == j:
                continue
            d = ((xi - xj) ** 2 + (yi - yj) ** 2) ** 0.5
            if d < nn_d:
                nn_d = d
                nn_s = scores[j]
        edge_d = float(min(xi, yi, W - xi - 1, H - yi - 1))
        feats[i] = [
            si,                                     # peak score
            float(heatmap[y0:y1, x0:x1].mean()),     # heat 5x5 mean
            float(heatmap[y0b:y1b, x0b:x1b].mean()), # heat 9x9 mean
            float(heatmap[y0:y1, x0:x1].std()),      # heat 5x5 std
            float(yellow[y0b:y1b, x0b:x1b].mean()),  # yellow 9x9 mean
            float(yellow[y0b:y1b, x0b:x1b].max()),   # yellow 9x9 max
            float(gray[y0b:y1b, x0b:x1b].mean()),    # gray 9x9 mean
            float(blue[y0b:y1b, x0b:x1b].mean()),    # blue 9x9 mean
            float(nn_d),                             # nearest neighbor distance
            float(nn_s),                             # nearest neighbor score
            edge_d,                                  # distance to image border
        ]
    return feats


def filter_false_positives(peaks: list[tuple[int, int]], heatmap: np.ndarray,
                           rgb: np.ndarray, classifier,
                           threshold: float = CLASSIFIER_THRESHOLD) -> list[tuple[int, int]]:
    """Drop peaks the classifier thinks are false positives."""
    if len(peaks) == 0 or classifier is None:
        return peaks
    feats = _peak_features(peaks, heatmap, rgb)
    probs = classifier.predict_proba(feats)[:, 1]
    return [p for p, prob in zip(peaks, probs) if prob >= threshold]


# ----------------------------------------------------------------------------
# Visualization
# ----------------------------------------------------------------------------

HEADER_HEIGHT = 260
LINE_COLOR = (255, 100, 0)   # BGR — orange
LINE_COLOR_5 = (0, 0, 255)   # BGR — red (top 5% emphasis)
LINE_THICKNESS = 2
DOT_COLOR = (0, 255, 0)      # BGR — green
DOT_RADIUS = 5


def render_annotated(img_bgr: np.ndarray, peaks: list[tuple[int, int]],
                     scan_name: str) -> tuple[np.ndarray, dict, list]:
    """Draw detections + split lines + header.

    Percentile lines are based on detected-pupa Y range:
        rank 0%   = pupa at image BOTTOM (largest y, physically highest in vial)
        rank 100% = pupa at image TOP    (smallest y, physically lowest in vial)
    Lines at rank 5%, 25%, 75%.

    Returns (annotated_img, counts_dict, per_pupa_list) where per_pupa_list
    is one dict per detected pupa with its exact rank percentile and band.
    """
    h, w = img_bgr.shape[:2]

    # Default stats dict (in case no pupa detected)
    counts = {
        "top_5_pct": 0,
        "rank_5_to_25_pct": 0,
        "middle_50_pct": 0,
        "bottom_25_pct": 0,
        "y_min_of_pupae": None,
        "y_max_of_pupae": None,
    }
    per_pupa = []

    if peaks:
        ys = np.array([y for x, y in peaks])
        y_min = int(ys.min())   # topmost pupa in image (smallest y) = rank 100
        y_max = int(ys.max())   # bottommost pupa in image (largest y) = rank 0
        y_range = max(1, y_max - y_min)
        counts["y_min_of_pupae"] = y_min
        counts["y_max_of_pupae"] = y_max

        # Line y for rank p (0 = image bottom, 100 = image top)
        def line_y(p: float) -> int:
            return int(y_max - (p / 100.0) * y_range)

        y_5  = line_y(5)
        y_25 = line_y(25)
        y_75 = line_y(75)

        for idx, (x, y) in enumerate(peaks):
            # rank_pct: 0 = image bottom (largest y) → 100 = image top (smallest y)
            rank_pct = (y_max - y) / y_range * 100.0
            if y >= y_5:
                band = "0-5%"
                counts["top_5_pct"] += 1
            elif y >= y_25:
                band = "5-25%"
                counts["rank_5_to_25_pct"] += 1
            elif y >= y_75:
                band = "25-75%"
                counts["middle_50_pct"] += 1
            else:
                band = "75-100%"
                counts["bottom_25_pct"] += 1
            per_pupa.append({
                "scan_name": scan_name,
                "pupa_idx": idx + 1,
                "x_pixel": int(x),
                "y_pixel": int(y),
                "rank_pct": round(rank_pct, 2),
                "band": band,
                "image_height": h,
            })
    else:
        y_5 = y_25 = y_75 = None

    # Draw on a copy of the scan
    marked = img_bgr.copy()
    for x, y in peaks:
        cv2.circle(marked, (x, y), DOT_RADIUS, DOT_COLOR, -1)

    if y_5 is not None:
        cv2.line(marked, (0, y_5),  (w, y_5),  LINE_COLOR_5, LINE_THICKNESS)
        cv2.line(marked, (0, y_25), (w, y_25), LINE_COLOR, LINE_THICKNESS)
        cv2.line(marked, (0, y_75), (w, y_75), LINE_COLOR, LINE_THICKNESS)
        cv2.putText(marked, "5% line",  (w - 140, y_5  - 8),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, LINE_COLOR_5, 1)
        cv2.putText(marked, "25% line", (w - 140, y_25 - 8),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, LINE_COLOR, 1)
        cv2.putText(marked, "75% line", (w - 140, y_75 - 8),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, LINE_COLOR, 1)

    # Header canvas
    final = np.ones((h + HEADER_HEIGHT, w, 3), dtype=np.uint8) * 255
    final[HEADER_HEIGHT:, :] = marked

    font = cv2.FONT_HERSHEY_SIMPLEX
    cv2.putText(final, scan_name, (20, 35), font, 0.8, (0, 0, 0), 2)
    cv2.putText(final, f"Total detected: {len(peaks)}", (20, 65),
                font, 0.65, (0, 0, 0), 2)
    cv2.putText(final, "Rank lines = percentiles of detected-pupa Y range",
                (20, 90), font, 0.45, (80, 80, 80), 1)
    cv2.putText(final, "(rank 0% = pupa at IMAGE BOTTOM ; rank 100% = pupa at IMAGE TOP)",
                (20, 110), font, 0.42, (80, 80, 80), 1)

    y_text = 140
    cv2.putText(final, f"Rank  0 - 5%   (top of ranking):       {counts['top_5_pct']}",
                (20, y_text), font, 0.55, LINE_COLOR_5, 2)
    cv2.putText(final, f"Rank  5 - 25%:                         {counts['rank_5_to_25_pct']}",
                (20, y_text + 30), font, 0.55, (0, 0, 0), 2)
    cv2.putText(final, f"Rank 25 - 75%  (middle 50%):           {counts['middle_50_pct']}",
                (20, y_text + 60), font, 0.55, (0, 0, 0), 2)
    cv2.putText(final, f"Rank 75 - 100% (bottom of ranking):    {counts['bottom_25_pct']}",
                (20, y_text + 90), font, 0.55, (0, 0, 0), 2)
    return final, counts, per_pupa


# ----------------------------------------------------------------------------
# Excel export
# ----------------------------------------------------------------------------

SUMMARY_SHEET = "Pupa counts"
DETAIL_SHEET = "Per pupa detail"

EXCEL_HEADERS = [
    "scan_name", "timestamp", "total_count",
    "top_5_pct_count",         # rank 0-5% (pupae closest to image BOTTOM)
    "rank_5_to_25_pct_count",  # rank 5-25%
    "middle_50_pct_count",     # rank 25-75%
    "bottom_25_pct_count",     # rank 75-100% (pupae closest to image TOP)
    "y_min_of_pupae", "y_max_of_pupae",
    "image_width", "image_height",
]

DETAIL_HEADERS = [
    "scan_name", "pupa_idx", "x_pixel", "y_pixel",
    "rank_pct",     # 0 = image bottom (largest y) … 100 = image top (smallest y)
    "band",
    "image_height",
]


def append_to_excel(excel_path: Path, scan_name: str, counts: dict,
                    total: int, img_w: int, img_h: int,
                    per_pupa: list) -> None:
    """Append one summary row + N per-pupa rows to the Excel file."""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    if excel_path.exists():
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # drop the default blank sheet

    # --- Summary sheet ---
    if SUMMARY_SHEET in wb.sheetnames:
        ws = wb[SUMMARY_SHEET]
    else:
        ws = wb.create_sheet(SUMMARY_SHEET, 0)
        ws.append(EXCEL_HEADERS)
    ws.append([
        scan_name,
        datetime.now().isoformat(timespec="seconds"),
        total,
        counts["top_5_pct"],
        counts["rank_5_to_25_pct"],
        counts["middle_50_pct"],
        counts["bottom_25_pct"],
        counts["y_min_of_pupae"],
        counts["y_max_of_pupae"],
        img_w,
        img_h,
    ])

    # --- Per-pupa detail sheet ---
    if DETAIL_SHEET in wb.sheetnames:
        ws2 = wb[DETAIL_SHEET]
    else:
        ws2 = wb.create_sheet(DETAIL_SHEET)
        ws2.append(DETAIL_HEADERS)
    for p in per_pupa:
        ws2.append([p["scan_name"], p["pupa_idx"], p["x_pixel"], p["y_pixel"],
                    p["rank_pct"], p["band"], p["image_height"]])

    wb.save(excel_path)


# ----------------------------------------------------------------------------
# Main driver
# ----------------------------------------------------------------------------

def process_one(model: TinyUNet, device: torch.device, src: Path,
                out_dir: Path, excel_path: Path, classifier=None) -> None:
    img_bgr = cv2.imread(str(src))
    if img_bgr is None:
        print(f"[skip] could not read {src}")
        return
    img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)

    heatmap = predict_heatmap(model, img_rgb, device)
    raw_peaks = extract_peaks(heatmap)

    # Optional 2nd-stage filter: drop peaks the classifier thinks are FPs
    if classifier is not None:
        peaks = filter_false_positives(raw_peaks, heatmap, img_rgb, classifier)
        n_filtered = len(raw_peaks) - len(peaks)
    else:
        peaks = raw_peaks
        n_filtered = 0

    annotated, counts, per_pupa = render_annotated(img_bgr, peaks, src.name)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_img = out_dir / f"{src.stem}_counted.png"
    cv2.imwrite(str(out_img), annotated)

    h, w = img_bgr.shape[:2]
    append_to_excel(excel_path, src.name, counts, len(peaks), w, h, per_pupa)

    filter_info = f"(filtered {n_filtered} FPs)" if n_filtered > 0 else ""
    print(f"[done] {src.name:40s} total={len(peaks):4d} {filter_info:>18}  "
          f"top5%={counts['top_5_pct']:3d}  "
          f"5-25%={counts['rank_5_to_25_pct']:3d}  "
          f"mid50%={counts['middle_50_pct']:3d}  "
          f"bot25%={counts['bottom_25_pct']:3d}  "
          f"-> {out_img.name}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Pupa counter — runs the v11 CNN on image(s), "
                    "saves annotated PNGs, and appends rows to an Excel file.",
    )
    parser.add_argument("input", type=Path,
                        help="Image file or directory of images (PNG/JPG).")
    parser.add_argument("--out", type=Path, default=Path("output"),
                        help="Directory for annotated PNGs (default: ./output)")
    parser.add_argument("--excel", type=Path, default=Path("output/pupa_counts.xlsx"),
                        help="Excel file to append counts to "
                             "(default: ./output/pupa_counts.xlsx)")
    parser.add_argument("--model", type=Path,
                        default=Path(__file__).resolve().parent / "model" / "pupa_counter_v12.pt",
                        help="Path to model weights (default: ./model/pupa_counter_v12.pt)")
    parser.add_argument("--no-filter", action="store_true",
                        help="Disable the peak-level FP classifier (useful for debugging).")
    parser.add_argument("--json-out", type=str, default=None,
                        help="Emit detection as JSON instead of PNG+Excel. "
                             "Use '-' for stdout. Skips all disk writes; "
                             "designed for the desktop app's subprocess call.")
    args = parser.parse_args()

    device = pick_device()
    print(f"Device: {device}")

    if not args.model.exists():
        print(f"ERROR: model weights not found at {args.model}")
        sys.exit(1)

    model = TinyUNet().to(device)
    model.load_state_dict(torch.load(args.model, map_location=device))
    model.eval()
    print(f"Loaded {args.model.name}")

    # Load classifier if present and not disabled
    classifier = None
    if not args.no_filter and CLASSIFIER_PATH.exists():
        import pickle
        with open(CLASSIFIER_PATH, "rb") as f:
            classifier = pickle.load(f)
        print(f"Loaded 2nd-stage filter: {CLASSIFIER_PATH.name}")
    elif args.no_filter:
        print("2nd-stage filter DISABLED (via --no-filter)")

    # Gather inputs
    if args.input.is_dir():
        images = sorted([p for p in args.input.iterdir()
                         if p.suffix.lower() in (".png", ".jpg", ".jpeg", ".tif", ".tiff")])
        if not images:
            print(f"No images found in {args.input}")
            sys.exit(1)
    elif args.input.is_file():
        images = [args.input]
    else:
        print(f"ERROR: {args.input} is not a file or directory")
        sys.exit(1)

    # --- JSON mode: skip disk writes, emit JSON for desktop-app subprocess --
    if args.json_out:
        import json, os, tempfile
        # ONE entry per requested image, in input order. Failed decodes
        # produce {"imagePath", "error"} records so the caller can align
        # responses to requests 1:1 and distinguish "no pupae" from
        # "image unreadable".
        results = []
        for img_path in images:
            try:
                img_bgr = cv2.imread(str(img_path))
                if img_bgr is None:
                    results.append({
                        "imagePath": str(img_path),
                        "error": "imread returned None (corrupt, unsupported format, or unreadable)",
                    })
                    continue
                img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
                heatmap = predict_heatmap(model, img_rgb, device)
                raw_peaks = extract_peaks(heatmap)
                if classifier is not None:
                    peaks = filter_false_positives(raw_peaks, heatmap, img_rgb, classifier)
                else:
                    peaks = raw_peaks
                _, counts, per_pupa = render_annotated(img_bgr, peaks, img_path.name)
                h, w = img_bgr.shape[:2]
                results.append({
                    "imagePath": str(img_path),
                    "imageWidth": w,
                    "imageHeight": h,
                    "modelVersion": args.model.name,
                    "pupae": [
                        {
                            "index": p["pupa_idx"],
                            "x": p["x_pixel"],
                            "y": p["y_pixel"],
                            "rankPct": p["rank_pct"],
                            "band": p["band"],
                            "source": "cnn",
                        }
                        for p in per_pupa
                    ],
                    "counts": {
                        "total": len(peaks),
                        "top5Pct": counts["top_5_pct"],
                        "rank5To25": counts["rank_5_to_25_pct"],
                        "middle50": counts["middle_50_pct"],
                        "bottom25": counts["bottom_25_pct"],
                    },
                    "yMin": counts["y_min_of_pupae"],
                    "yMax": counts["y_max_of_pupae"],
                })
            except Exception as exc:
                import traceback as _tb
                results.append({
                    "imagePath": str(img_path),
                    "error": f"{type(exc).__name__}: {exc}",
                    "traceback": _tb.format_exc(),
                })
        # Top-level shape is driven by whether the user passed a FILE or a
        # DIRECTORY, not by how many succeeded. Callers can rely on this.
        if args.input.is_file():
            payload = results[0]
        else:
            payload = results
        text = json.dumps(payload, separators=(",", ":"))

        # --- Writing JSON to stdout or a file ---
        if args.json_out == "-":
            # Sentinel + JSON — makes it trivial for a subprocess reader to
            # skip any unrelated stdout noise (e.g. "Device: mps").
            print("<<<JSON>>>")
            print(text)
            print("<<<END>>>")
        else:
            out_path = Path(args.json_out).resolve()
            # Refuse to overwrite any input image — a caller-side path bug
            # must not silently truncate a scan.
            input_resolved = {Path(p).resolve() for p in images}
            if out_path in input_resolved:
                print(f"ERROR: --json-out resolves to an input image, refusing: {out_path}",
                      file=sys.stderr)
                sys.exit(3)
            # Atomic write: tmp file in the same directory, then rename.
            # Avoids half-written JSON if the process dies mid-write.
            out_path.parent.mkdir(parents=True, exist_ok=True)
            fd, tmp_name = tempfile.mkstemp(
                prefix=out_path.stem + ".", suffix=".tmp.json",
                dir=str(out_path.parent),
            )
            try:
                with os.fdopen(fd, "w") as tmpf:
                    tmpf.write(text)
                os.replace(tmp_name, out_path)
            except Exception:
                try:
                    os.unlink(tmp_name)
                except OSError:
                    pass
                raise
        return

    print(f"Processing {len(images)} image(s), writing to {args.excel}\n")
    args.excel.parent.mkdir(parents=True, exist_ok=True)

    for img_path in images:
        process_one(model, device, img_path, args.out, args.excel, classifier=classifier)

    # Distribution plot: one global histogram + per-scan small multiples
    try:
        plot_path = args.out / "rank_distribution.png"
        generate_distribution_plot(args.excel, plot_path)
        print(f"Distribution plot: {plot_path}")
    except Exception as exc:
        print(f"(skipped distribution plot: {exc})")

    print(f"\nDone. Annotated images: {args.out}")
    print(f"Excel log: {args.excel}")


def generate_distribution_plot(excel_path: Path, out_png: Path) -> None:
    """Read the per-pupa detail sheet and produce a 2-panel matplotlib figure:
    (left) global rank-percentile histogram across all pupae;
    (right) small multiples: one histogram per scan."""
    from openpyxl import load_workbook
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    if DETAIL_SHEET not in wb.sheetnames:
        raise RuntimeError(f"sheet '{DETAIL_SHEET}' not found")
    ws = wb[DETAIL_SHEET]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise RuntimeError("no per-pupa rows to plot")

    from collections import defaultdict
    all_ranks = []
    by_scan = defaultdict(list)
    for scan_name, pupa_idx, x_px, y_px, rank_pct, band, img_h in rows:
        if rank_pct is None:
            continue
        all_ranks.append(float(rank_pct))
        by_scan[scan_name].append(float(rank_pct))

    scans = sorted(by_scan.keys())
    n_scans = len(scans)
    # Grid layout: at most 5 columns, grow rows as needed
    cols = min(5, max(1, n_scans))
    rows_grid = (n_scans + cols - 1) // cols

    fig = plt.figure(figsize=(14, 4 + 1.8 * rows_grid))
    gs = fig.add_gridspec(1 + rows_grid, cols, height_ratios=[2.2] + [1] * rows_grid)

    # Global histogram
    ax = fig.add_subplot(gs[0, :])
    ax.hist(all_ranks, bins=20, range=(0, 100), color="steelblue", edgecolor="white")
    ax.axvline(5, color="red", linestyle="--", linewidth=1, label="5%")
    ax.axvline(25, color="orange", linestyle="--", linewidth=1, label="25%")
    ax.axvline(75, color="orange", linestyle="--", linewidth=1, label="75%")
    ax.set_xlabel("Rank percentile (0 = image bottom  →  100 = image top)")
    ax.set_ylabel("Pupa count")
    ax.set_title(f"Rank distribution across all {len(all_ranks)} pupae "
                 f"in {n_scans} scans")
    ax.legend(loc="upper right", fontsize=8)
    ax.set_xlim(0, 100)

    # Per-scan small multiples
    for i, scan_name in enumerate(scans):
        r = 1 + i // cols
        c = i % cols
        axi = fig.add_subplot(gs[r, c])
        axi.hist(by_scan[scan_name], bins=15, range=(0, 100),
                 color="steelblue", edgecolor="white")
        axi.set_title(scan_name.replace("Scan_20260313 ", "").replace(".png", ""),
                      fontsize=8)
        axi.set_xlim(0, 100)
        axi.tick_params(axis="both", labelsize=7)

    fig.tight_layout()
    fig.savefig(out_png, dpi=120)
    plt.close(fig)


if __name__ == "__main__":
    main()
